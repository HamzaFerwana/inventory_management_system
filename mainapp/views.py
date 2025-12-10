from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from .models import (
    Category,
    SubCategory,
    Product,
    Customer,
    ExpenseCategory,
    Expense,
    Supplier,
    Quotation,
    Purchase,
)
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from decimal import Decimal, ROUND_HALF_UP
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from django.db import transaction


MONEY_QUANT = Decimal("0.01")


def compute_purchase_totals(quantity, unit_price, discount_pct, tax_pct):

    subtotal_base = unit_price * quantity

    discount_amount = subtotal_base * (discount_pct / Decimal("100"))

    taxable_base = subtotal_base
    tax_amount = taxable_base * (tax_pct / Decimal("100"))

    line_total = subtotal_base - discount_amount + tax_amount

    discount_amount = discount_amount.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)
    tax_amount = tax_amount.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)
    line_total = line_total.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)

    return discount_amount, tax_amount, line_total


@login_required
def index(request):
    return render(request, "index.html")


@login_required
def category_create(request):
    errors = []
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        code = request.POST.get("code", "").strip()
        if not name:
            errors.append("Category name is required.")
        if not code:
            errors.append("Category code is required.")
        if Category.objects.filter(code=code).exists():
            errors.append("Category code must be unique.")
        if not errors:
            Category.objects.create(
                name=name,
                code=code,
            )
            messages.success(request, "Category created successfully!")
            return redirect("categories_index")
    context = {
        "errors": errors,
        "old": request.POST if request.method == "POST" else {},
    }
    return render(request, "addcategory.html", context)


@login_required
def categories_index(request):
    categories = Category.objects.all().order_by("name")
    return render(request, "categorylist.html", {"categories": categories})


@login_required
def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    errors = []
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        code = request.POST.get("code", "").strip()
        if not name:
            errors.append("Category name is required.")
        if not code:
            errors.append("Category code is required.")
        if code and Category.objects.filter(code=code).exclude(pk=category.pk).exists():
            errors.append("Category code must be unique.")
        if not errors:
            category.name = name
            category.code = code
            category.save()
            messages.success(request, "Category updated successfully!")
            return redirect("categories_index")
        old = {"name": name, "code": code}
    else:
        old = {"name": category.name, "code": category.code}
    context = {
        "errors": errors,
        "old": old,
        "category": category,
    }
    return render(request, "editcategory.html", context)


@login_required
def categories_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        category.delete()
        messages.success(request, "Category deleted successfully.")
        return redirect("categories_index")
    category.delete()
    messages.success(request, "Category deleted successfully.")
    return redirect("categories_index")


@login_required
def subcategories_index(request):
    subcategories = SubCategory.objects.select_related("category").order_by("name")
    return render(
        request,
        "subcategorylist.html",
        {"subcategories": subcategories},
    )


@login_required
def subcategory_create(request):
    errors = []

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        code = request.POST.get("code", "").strip()
        description = request.POST.get("description", "").strip()
        category_id = request.POST.get("category_id")

        if not name:
            errors.append("Subcategory name is required.")
        if not code:
            errors.append("Subcategory code is required.")
        if not category_id:
            errors.append("Category is required.")

        if code and SubCategory.objects.filter(code=code).exists():
            errors.append("Subcategory code must be unique.")

        category = None
        if category_id:
            category = Category.objects.filter(pk=category_id).first()
            if not category:
                errors.append("Selected category does not exist.")

        if not errors:
            SubCategory.objects.create(
                name=name,
                code=code,
                description=description or None,
                category=category,
            )
            messages.success(request, "Subcategory created successfully!")
            return redirect("subcategories_index")

        old = {
            "name": name,
            "code": code,
            "description": description,
            "category_id": category_id,
        }
    else:
        old = {}

    context = {
        "errors": errors,
        "old": old,
        "categories": Category.objects.all().order_by("name"),
    }
    return render(request, "addsubcategory.html", context)


@login_required
def subcategory_edit(request, pk):
    subcategory = get_object_or_404(SubCategory, pk=pk)
    errors = []

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        code = request.POST.get("code", "").strip()
        description = request.POST.get("description", "").strip()
        category_id = request.POST.get("category_id")

        if not name:
            errors.append("Subcategory name is required.")
        if not code:
            errors.append("Subcategory code is required.")
        if not category_id:
            errors.append("Category is required.")

        if (
            code
            and SubCategory.objects.filter(code=code)
            .exclude(pk=subcategory.pk)
            .exists()
        ):
            errors.append("Subcategory code must be unique.")

        category = None
        if category_id:
            category = Category.objects.filter(pk=category_id).first()
            if not category:
                errors.append("Selected category does not exist.")

        if not errors:
            subcategory.name = name
            subcategory.code = code
            subcategory.description = description or None
            subcategory.category = category
            subcategory.save()
            messages.success(request, "Subcategory updated successfully!")
            return redirect("subcategories_index")

        old = {
            "name": name,
            "code": code,
            "description": description,
            "category_id": category_id,
        }
    else:
        old = {
            "name": subcategory.name,
            "code": subcategory.code,
            "description": subcategory.description or "",
            "category_id": subcategory.category_id,
        }

    context = {
        "errors": errors,
        "old": old,
        "subcategory": subcategory,
        "categories": Category.objects.all().order_by("name"),
    }
    return render(request, "editsubcategory.html", context)


@login_required
def subcategories_delete(request, pk):
    subcategory = get_object_or_404(SubCategory, pk=pk)

    if request.method == "POST":
        subcategory.delete()
        messages.success(request, "Subcategory deleted successfully.")
        return redirect("subcategories_index")

    subcategory.delete()
    messages.success(request, "Subcategory deleted successfully.")
    return redirect("subcategories_index")


@login_required
def products_index(request):
    products = Product.objects.select_related("sub_category").order_by("name")

    for p in products:
        price = p.price or 0
        disc = p.discount_percentage or 0
        p.discounted_price = price * (Decimal("1") - Decimal(disc) / Decimal("100"))

    return render(request, "productlist.html", {"products": products})


@login_required
def product_create(request):
    errors = []

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        sub_category_id = request.POST.get("sub_category_id")
        unit = request.POST.get("unit", "").strip()
        sku = request.POST.get("sku", "").strip()
        quantity = request.POST.get("quantity", "0").strip()
        description = request.POST.get("description", "").strip()
        status = request.POST.get("status", "active")
        price = request.POST.get("price", "").strip()
        discount_percentage = request.POST.get("discount_percentage", "0").strip()
        main_image = request.FILES.get("main_image")

        if not name:
            errors.append("Product name is required.")
        if not unit:
            errors.append("Unit is required.")
        if not sku:
            errors.append("SKU is required.")
        if sku and Product.objects.filter(sku=sku).exists():
            errors.append("SKU must be unique.")
        if not price:
            errors.append("Price is required.")

        sub_category = None
        if sub_category_id:
            sub_category = SubCategory.objects.filter(pk=sub_category_id).first()
            if not sub_category:
                errors.append("Selected subcategory does not exist.")

        try:
            quantity_val = int(quantity or 0)
            if quantity_val < 0:
                errors.append("Quantity cannot be negative.")
        except ValueError:
            errors.append("Quantity must be a number.")

        if not errors:
            Product.objects.create(
                name=name,
                sub_category=sub_category,
                unit=unit,
                sku=sku,
                quantity=quantity_val,
                description=description or None,
                status=status,
                price=price or 0,
                discount_percentage=discount_percentage or 0,
                main_image=main_image,
            )
            messages.success(request, "Product created successfully!")
            return redirect("products_index")

        old = {
            "name": name,
            "sub_category_id": sub_category_id,
            "unit": unit,
            "sku": sku,
            "quantity": quantity,
            "description": description,
            "status": status,
            "price": price,
            "discount_percentage": discount_percentage,
        }
    else:
        old = {}

    context = {
        "errors": errors,
        "old": old,
        "subcategories": SubCategory.objects.select_related("category").order_by(
            "category__name", "name"
        ),
    }
    return render(request, "addproduct.html", context)


@login_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    errors = []

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        sub_category_id = request.POST.get("sub_category_id")
        unit = request.POST.get("unit", "").strip()
        sku = request.POST.get("sku", "").strip()
        quantity = request.POST.get("quantity", "0").strip()
        description = request.POST.get("description", "").strip()
        status = request.POST.get("status", "active")
        price = request.POST.get("price", "").strip()
        discount_percentage = request.POST.get("discount_percentage", "0").strip()
        main_image = request.FILES.get("main_image")

        if not name:
            errors.append("Product name is required.")
        if not unit:
            errors.append("Unit is required.")
        if not sku:
            errors.append("SKU is required.")
        if sku and Product.objects.filter(sku=sku).exclude(pk=product.pk).exists():
            errors.append("SKU must be unique.")
        if not price:
            errors.append("Price is required.")

        sub_category = None
        if sub_category_id:
            sub_category = SubCategory.objects.filter(pk=sub_category_id).first()
            if not sub_category:
                errors.append("Selected subcategory does not exist.")

        try:
            quantity_val = int(quantity or 0)
            if quantity_val < 0:
                errors.append("Quantity cannot be negative.")
        except ValueError:
            errors.append("Quantity must be a number.")

        if not errors:
            product.name = name
            product.sub_category = sub_category
            product.unit = unit
            product.sku = sku
            product.quantity = quantity_val
            product.description = description or None
            product.status = status
            product.price = price or 0
            product.discount_percentage = discount_percentage or 0
            if main_image:
                product.main_image = main_image
            product.save()
            messages.success(request, "Product updated successfully!")
            return redirect("products_index")

        old = {
            "name": name,
            "sub_category_id": sub_category_id,
            "unit": unit,
            "sku": sku,
            "quantity": quantity,
            "description": description,
            "status": status,
            "price": price,
            "discount_percentage": discount_percentage,
        }
    else:
        old = {
            "name": product.name,
            "sub_category_id": product.sub_category_id,
            "unit": product.unit,
            "sku": product.sku,
            "quantity": product.quantity,
            "description": product.description or "",
            "status": product.status,
            "price": product.price,
            "discount_percentage": product.discount_percentage,
        }

    context = {
        "errors": errors,
        "old": old,
        "product": product,
        "subcategories": SubCategory.objects.select_related("category").order_by(
            "category__name", "name"
        ),
    }
    return render(request, "editproduct.html", context)


@login_required
def products_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)

    if request.method == "POST":
        product.delete()
        messages.success(request, "Product deleted successfully.")
        return redirect("products_index")

    product.delete()
    messages.success(request, "Product deleted successfully.")
    return redirect("products_index")


def register(request):
    if request.user.is_authenticated:
        return redirect("index")

    errors = []
    username = ""
    email = ""

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "")

        if not username or not email or not password:
            errors.append("All fields are required.")

        if User.objects.filter(username=username).exists():
            errors.append("This username is already taken.")

        if User.objects.filter(email=email).exists():
            errors.append("This email is already in use.")

        if len(password) < 8:
            errors.append("Password must be at least 8 characters long.")

        if not errors:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
            )
            login(request, user)
            return redirect("index")

    context = {
        "errors": errors,
        "username": username,
        "email": email,
    }
    return render(request, "register.html", context)


def login_view(request):
    if request.user.is_authenticated:
        return redirect("index")

    error = None

    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "")

        user = None
        if not email or not password:
            error = "Email and password are required."
        else:
            try:
                u = User.objects.get(email=email)
                user = authenticate(request, username=u.username, password=password)
            except User.DoesNotExist:
                user = None

        if user is not None:
            login(request, user)
            return redirect("index")
        else:
            if error is None:
                error = "Invalid email or password."
            return render(request, "login.html", {"error": error})

    return render(request, "login.html")


@login_required
def logout_view(request):
    if request.method == "POST":
        logout(request)
        return redirect("login")
    return redirect("index")


@login_required
def customers_index(request):
    customers = Customer.objects.all().order_by("name")
    return render(request, "customerlist.html", {"customers": customers})


@login_required
def customer_create(request):
    errors = []
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        email = request.POST.get("email", "").strip()
        phone = request.POST.get("phone", "").strip()
        country = request.POST.get("country", "").strip()
        city = request.POST.get("city", "").strip()
        address = request.POST.get("address", "").strip()
        description = request.POST.get("description", "").strip()
        avatar = request.FILES.get("avatar")

        if not name:
            errors.append("Customer name is required.")
        if not email:
            errors.append("Email is required.")
        if not phone:
            errors.append("Phone is required.")

        if email and Customer.objects.filter(email=email).exists():
            errors.append("This email is already in use.")

        if not errors:
            Customer.objects.create(
                name=name,
                email=email,
                phone=phone,
                country=country or None,
                city=city or None,
                address=address or None,
                description=description or None,
                avatar=avatar,
            )
            messages.success(request, "Customer created successfully!")
            return redirect("customers_index")

        old = {
            "name": name,
            "email": email,
            "phone": phone,
            "country": country,
            "city": city,
            "address": address,
            "description": description,
        }
    else:
        old = {}

    context = {
        "errors": errors,
        "old": old,
    }
    return render(request, "addcustomer.html", context)


@login_required
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    errors = []

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        email = request.POST.get("email", "").strip()
        phone = request.POST.get("phone", "").strip()
        country = request.POST.get("country", "").strip()
        city = request.POST.get("city", "").strip()
        address = request.POST.get("address", "").strip()
        description = request.POST.get("description", "").strip()
        avatar = request.FILES.get("avatar")

        if not name:
            errors.append("Customer name is required.")
        if not email:
            errors.append("Email is required.")
        if not phone:
            errors.append("Phone is required.")

        if (
            email
            and Customer.objects.filter(email=email).exclude(pk=customer.pk).exists()
        ):
            errors.append("This email is already in use.")

        if not errors:
            customer.name = name
            customer.email = email
            customer.phone = phone
            customer.country = country or None
            customer.city = city or None
            customer.address = address or None
            customer.description = description or None
            if avatar:
                customer.avatar = avatar
            customer.save()
            messages.success(request, "Customer updated successfully!")
            return redirect("customers_index")

        old = {
            "name": name,
            "email": email,
            "phone": phone,
            "country": country,
            "city": city,
            "address": address,
            "description": description,
        }
    else:
        old = {
            "name": customer.name,
            "email": customer.email,
            "phone": customer.phone,
            "country": customer.country or "",
            "city": customer.city or "",
            "address": customer.address or "",
            "description": customer.description or "",
        }

    context = {
        "errors": errors,
        "old": old,
        "customer": customer,
    }
    return render(request, "editcustomer.html", context)


@login_required
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)

    if request.method == "POST":
        customer.delete()
        messages.success(request, "Customer deleted successfully.")
        return redirect("customers_index")

    customer.delete()
    messages.success(request, "Customer deleted successfully.")
    return redirect("customers_index")


class DataExporter:
    @staticmethod
    def export_to_pdf(queryset, fields, title, filename_prefix):
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        )

        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=10,
            textColor=colors.HexColor("#2c3e50"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.3 * inch))

        headers = [display_name for _, display_name in fields]
        data = [headers]

        for obj in queryset:
            row = []
            for field_name, _ in fields:

                value = obj
                for attr in field_name.split("."):
                    value = getattr(value, attr, "")
                    if value is None:
                        value = ""
                row.append(str(value))
            data.append(row)
        num_cols = len(fields)
        available_width = 6.5 * inch
        col_width = available_width / num_cols
        table = Table(data, colWidths=[col_width] * num_cols)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 4),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.lightgrey],
                    ),
                ]
            )
        )

        elements.append(table)
        elements.append(Spacer(1, 0.5 * inch))
        footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {queryset.count()}"
        elements.append(Paragraph(footer_text, styles["Normal"]))

        doc.build(elements)
        return response

    @staticmethod
    def export_to_excel(queryset, fields, title, filename_prefix):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]

        header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4CAF50", end_color="4CAF50", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        num_cols = len(fields)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(name="Calibri", size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_num, (_, display_name) in enumerate(fields, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = display_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for row_num, obj in enumerate(queryset, 4):
            for col_num, (field_name, _) in enumerate(fields, 1):
                value = obj
                for attr in field_name.split("."):
                    value = getattr(value, attr, "")
                    if value is None:
                        value = ""

                cell = ws.cell(row=row_num, column=col_num)
                cell.value = str(value)
                cell.alignment = cell_alignment
                cell.border = border
        for col_num in range(1, num_cols + 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        last_row = ws.max_row + 2
        ws.cell(
            row=last_row,
            column=1,
            value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {queryset.count()}",
        )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        )

        wb.save(response)
        return response


@login_required
def export_categories_pdf(request):

    queryset = Category.objects.all().order_by("-created_at")
    fields = [
        ("id", "Category ID"),
        ("name", "Category Name"),
        ("code", "Category Code"),
    ]

    return DataExporter.export_to_pdf(
        queryset=queryset,
        fields=fields,
        title="Categories Report",
        filename_prefix="categories",
    )


@login_required
def export_categories_excel(request):
    queryset = Category.objects.all().order_by("-created_at")
    fields = [
        ("id", "Category ID"),
        ("name", "Category Name"),
        ("code", "Category Code"),
    ]

    return DataExporter.export_to_excel(
        queryset=queryset,
        fields=fields,
        title="Categories",
        filename_prefix="categories",
    )


@login_required
def export_products_pdf(request):

    queryset = (
        Product.objects.select_related("sub_category", "sub_category__category")
        .all()
        .order_by("-created_at")
    )
    fields = [
        ("id", "Product ID"),
        ("name", "Product Name"),
        ("sku", "SKU"),
        ("sub_category.category.name", "Category"),
        ("sub_category.name", "Sub Category"),
        ("unit", "Unit"),
        ("quantity", "Quantity"),
        ("price", "Price"),
        ("discount_percentage", "Discount %"),
        ("status", "Status"),
    ]

    return DataExporter.export_to_pdf(
        queryset=queryset,
        fields=fields,
        title="Products Report",
        filename_prefix="products",
    )


@login_required
def export_products_excel(request):
    queryset = (
        Product.objects.select_related("sub_category", "sub_category__category")
        .all()
        .order_by("-created_at")
    )
    fields = [
        ("id", "Product ID"),
        ("name", "Product Name"),
        ("sku", "SKU"),
        ("sub_category.category.name", "Category"),
        ("sub_category.name", "Sub Category"),
        ("unit", "Unit"),
        ("quantity", "Quantity"),
        ("price", "Price"),
        ("discount_percentage", "Discount %"),
        ("status", "Status"),
        ("description", "Description"),
    ]
    return DataExporter.export_to_excel(
        queryset=queryset, fields=fields, title="Products", filename_prefix="products"
    )


@login_required
def export_subcategories_pdf(request):
    queryset = (
        SubCategory.objects.select_related("category").all().order_by("-created_at")
    )
    fields = [
        ("id", "Sub Category ID"),
        ("name", "Sub Category Name"),
        ("code", "Sub Category Code"),
        ("category.name", "Parent Category"),
        ("description", "Description"),
    ]

    return DataExporter.export_to_pdf(
        queryset=queryset,
        fields=fields,
        title="Sub Categories Report",
        filename_prefix="subcategories",
    )


@login_required
def export_subcategories_excel(request):

    queryset = (
        SubCategory.objects.select_related("category").all().order_by("-created_at")
    )
    fields = [
        ("id", "Sub Category ID"),
        ("name", "Sub Category Name"),
        ("code", "Sub Category Code"),
        ("category.name", "Parent Category"),
        ("description", "Description"),
    ]

    return DataExporter.export_to_excel(
        queryset=queryset,
        fields=fields,
        title="Sub Categories",
        filename_prefix="subcategories",
    )


@login_required
def expense_categories_index(request):
    categories = ExpenseCategory.objects.all()
    return render(
        request,
        "expense_categories/list.html",
        {
            "categories": categories,
        },
    )


@login_required
def expense_category_create(request):
    errors = []
    old = {}

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        code = request.POST.get("code", "").strip() or None
        description = request.POST.get("description", "").strip() or None

        if not name:
            errors.append("Name is required.")

        if code and ExpenseCategory.objects.filter(code=code).exists():
            errors.append("Code must be unique.")

        if not errors:
            ExpenseCategory.objects.create(
                name=name,
                code=code,
                description=description,
            )
            messages.success(request, "Expense category created successfully!")
            return redirect("expense_categories_index")

        old = {
            "name": name,
            "code": code or "",
            "description": description or "",
        }

    context = {
        "errors": errors,
        "old": old,
    }
    return render(request, "expense_categories/add.html", context)


@login_required
def expense_category_edit(request, pk):
    category = get_object_or_404(ExpenseCategory, pk=pk)
    errors = []

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        code = request.POST.get("code", "").strip() or None
        description = request.POST.get("description", "").strip() or None

        if not name:
            errors.append("Name is required.")

        if (
            code
            and ExpenseCategory.objects.filter(code=code)
            .exclude(pk=category.pk)
            .exists()
        ):
            errors.append("Code must be unique.")

        if not errors:
            category.name = name
            category.code = code
            category.description = description
            category.save()
            messages.success(request, "Expense category updated successfully!")
            return redirect("expense_categories_index")

        old = {
            "name": name,
            "code": code or "",
            "description": description or "",
        }
    else:
        old = {
            "name": category.name,
            "code": category.code or "",
            "description": category.description or "",
        }

    context = {
        "errors": errors,
        "old": old,
        "category": category,
    }
    return render(request, "expense_categories/edit.html", context)


@login_required
def expense_category_delete(request, pk):
    category = get_object_or_404(ExpenseCategory, pk=pk)

    if request.method == "POST":
        category.delete()
        messages.success(request, "Expense category deleted successfully.")
        return redirect("expense_categories_index")
    category.delete()
    messages.success(request, "Expense category deleted successfully.")
    return redirect("expense_categories_index")


@login_required
def expenses_index(request):
    expenses = Expense.objects.select_related("expense_category").order_by(
        "-date", "-id"
    )
    return render(request, "expenses/list.html", {"expenses": expenses})


@login_required
def expense_create(request):
    errors = []
    old = {}

    if request.method == "POST":
        expense_category_id = request.POST.get("expense_category_id")
        date = request.POST.get("date", "").strip()
        amount = request.POST.get("amount", "").strip()
        reference = request.POST.get("reference", "").strip()
        expense_for = request.POST.get("expense_for", "").strip()
        description = request.POST.get("description", "").strip()

        if not expense_category_id:
            errors.append("Category is required.")
        if not date:
            errors.append("Date is required.")
        if not amount:
            errors.append("Amount is required.")
        if not reference:
            errors.append("Reference is required.")
        if reference and Expense.objects.filter(reference=reference).exists():
            errors.append("Reference must be unique.")
        if not expense_for:
            errors.append("Expense for is required.")

        expense_category = None
        if expense_category_id:
            expense_category = ExpenseCategory.objects.filter(
                pk=expense_category_id
            ).first()
            if not expense_category:
                errors.append("Selected category does not exist.")

        amount_val = None
        if amount:
            try:
                amount_val = Decimal(amount)
                if amount_val <= 0:
                    errors.append("Amount must be greater than zero.")
            except Exception:
                errors.append("Amount must be a valid number.")

        if not errors:
            Expense.objects.create(
                expense_category=expense_category,
                date=date,
                amount=amount_val,
                reference=reference,
                expense_for=expense_for,
                description=description or None,
            )
            messages.success(request, "Expense created successfully!")
            return redirect("expenses_index")

        old = {
            "expense_category_id": expense_category_id,
            "date": date,
            "amount": amount,
            "reference": reference,
            "expense_for": expense_for,
            "description": description,
        }

    categories = ExpenseCategory.objects.order_by("name")
    context = {
        "errors": errors,
        "old": old,
        "categories": categories,
    }
    return render(request, "expenses/add.html", context)


@login_required
def expense_edit(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    errors = []

    if request.method == "POST":
        expense_category_id = request.POST.get("expense_category_id")
        date = request.POST.get("date", "").strip()
        amount = request.POST.get("amount", "").strip()
        reference = request.POST.get("reference", "").strip()
        expense_for = request.POST.get("expense_for", "").strip()
        description = request.POST.get("description", "").strip()

        if not expense_category_id:
            errors.append("Category is required.")
        if not date:
            errors.append("Date is required.")
        if not amount:
            errors.append("Amount is required.")
        if not reference:
            errors.append("Reference is required.")
        if (
            reference
            and Expense.objects.filter(reference=reference)
            .exclude(pk=expense.pk)
            .exists()
        ):
            errors.append("Reference must be unique.")
        if not expense_for:
            errors.append("Expense for is required.")

        expense_category = None
        if expense_category_id:
            expense_category = ExpenseCategory.objects.filter(
                pk=expense_category_id
            ).first()
            if not expense_category:
                errors.append("Selected category does not exist.")

        amount_val = None
        if amount:
            try:
                amount_val = Decimal(amount)
                if amount_val <= 0:
                    errors.append("Amount must be greater than zero.")
            except Exception:
                errors.append("Amount must be a valid number.")

        if not errors:
            expense.expense_category = expense_category
            expense.date = date
            expense.amount = amount_val
            expense.reference = reference
            expense.expense_for = expense_for
            expense.description = description or None
            expense.save()
            messages.success(request, "Expense updated successfully!")
            return redirect("expenses_index")

        old = {
            "expense_category_id": expense_category_id,
            "date": date,
            "amount": amount,
            "reference": reference,
            "expense_for": expense_for,
            "description": description,
        }
    else:
        old = {
            "expense_category_id": expense.expense_category_id,
            "date": expense.date,
            "amount": expense.amount,
            "reference": expense.reference,
            "expense_for": expense.expense_for,
            "description": expense.description or "",
        }

    categories = ExpenseCategory.objects.order_by("name")
    context = {
        "errors": errors,
        "old": old,
        "expense": expense,
        "categories": categories,
    }
    return render(request, "expenses/edit.html", context)


@login_required
def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)

    if request.method == "POST":
        expense.delete()
        messages.success(request, "Expense deleted successfully.")
        return redirect("expenses_index")

    expense.delete()
    messages.success(request, "Expense deleted successfully.")
    return redirect("expenses_index")


@login_required
def suppliers_index(request):

    suppliers = Supplier.objects.order_by("name")

    return render(request, "suppliers/list.html", {"suppliers": suppliers})


@login_required
def supplier_create(request):

    errors = []

    old = {}

    if request.method == "POST":

        name = request.POST.get("name", "").strip()

        email = request.POST.get("email", "").strip()

        phone = request.POST.get("phone", "").strip()

        country = request.POST.get("country", "").strip()

        city = request.POST.get("city", "").strip()

        address = request.POST.get("address", "").strip()

        description = request.POST.get("description", "").strip()

        avatar = request.FILES.get("avatar")

        if not name:

            errors.append("Name is required.")

        if email and Supplier.objects.filter(email=email).exists():

            errors.append("Email is already in use.")

        if not errors:

            Supplier.objects.create(
                name=name,
                email=email or None,
                phone=phone or None,
                country=country or None,
                city=city or None,
                address=address or None,
                description=description or None,
                avatar=avatar,
            )

            messages.success(request, "Supplier created successfully!")

            return redirect("suppliers_index")

        old = {
            "name": name,
            "email": email,
            "phone": phone,
            "country": country,
            "city": city,
            "address": address,
            "description": description,
        }

    context = {
        "errors": errors,
        "old": old,
    }

    return render(request, "suppliers/add.html", context)


@login_required
def supplier_edit(request, pk):

    supplier = get_object_or_404(Supplier, pk=pk)

    errors = []

    if request.method == "POST":

        name = request.POST.get("name", "").strip()

        email = request.POST.get("email", "").strip()

        phone = request.POST.get("phone", "").strip()

        country = request.POST.get("country", "").strip()

        city = request.POST.get("city", "").strip()

        address = request.POST.get("address", "").strip()

        description = request.POST.get("description", "").strip()

        avatar = request.FILES.get("avatar")

        if not name:

            errors.append("Name is required.")

        if (
            email
            and Supplier.objects.filter(email=email).exclude(pk=supplier.pk).exists()
        ):

            errors.append("Email is already in use.")

        if not errors:

            supplier.name = name

            supplier.email = email or None

            supplier.phone = phone or None

            supplier.country = country or None

            supplier.city = city or None

            supplier.address = address or None

            supplier.description = description or None

            if avatar:

                supplier.avatar = avatar

            supplier.save()

            messages.success(request, "Supplier updated successfully!")

            return redirect("suppliers_index")

        old = {
            "name": name,
            "email": email,
            "phone": phone,
            "country": country,
            "city": city,
            "address": address,
            "description": description,
        }

    else:

        old = {
            "name": supplier.name,
            "email": supplier.email or "",
            "phone": supplier.phone or "",
            "country": supplier.country or "",
            "city": supplier.city or "",
            "address": supplier.address or "",
            "description": supplier.description or "",
        }

    context = {
        "errors": errors,
        "old": old,
        "supplier": supplier,
    }

    return render(request, "suppliers/edit.html", context)


@login_required
def supplier_delete(request, pk):

    supplier = get_object_or_404(Supplier, pk=pk)

    if request.method == "POST":

        supplier.delete()

        messages.success(request, "Supplier deleted successfully.")

        return redirect("suppliers_index")

    supplier.delete()

    messages.success(request, "Supplier deleted successfully.")

    return redirect("suppliers_index")


@login_required
def quotations_index(request):
    quotations = Quotation.objects.select_related("product", "customer").order_by(
        "-created_at"
    )
    return render(request, "quotations/list.html", {"quotations": quotations})


@login_required
def quotation_create(request):
    products = Product.objects.order_by("name")
    customers = Customer.objects.order_by("name")
    errors = []

    if request.method == "POST":
        reference = request.POST.get("reference", "").strip()
        product_id = request.POST.get("product_id")
        customer_id = request.POST.get("customer_id")
        quantity_raw = request.POST.get("quantity", "1").strip()
        unit_price_raw = request.POST.get("unit_price", "0").strip()
        discount_raw = request.POST.get("discount_percentage", "0").strip()
        tax_raw = request.POST.get("tax_percentage", "0").strip()
        status = request.POST.get("status", "pending").strip()
        notes = request.POST.get("notes", "").strip()

        product = Product.objects.filter(pk=product_id).first()
        customer = Customer.objects.filter(pk=customer_id).first()

        try:
            quantity = int(quantity_raw)
            if quantity <= 0:
                raise ValueError
        except ValueError:
            errors.append("Quantity must be a positive integer.")
            quantity = None

        try:
            unit_price = Decimal(unit_price_raw)
        except Exception:
            errors.append("Unit price must be a number.")
            unit_price = None

        try:
            discount_percentage = Decimal(discount_raw or 0)
        except Exception:
            errors.append("Discount must be a number.")
            discount_percentage = None

        try:
            tax_percentage = Decimal(tax_raw or 0)
        except Exception:
            errors.append("Tax must be a number.")
            tax_percentage = None

        if not reference:
            errors.append("Reference is required.")
        elif Quotation.objects.filter(reference=reference).exists():
            errors.append("Reference must be unique.")

        if not product:
            errors.append("Product is required.")

        if not customer:
            errors.append("Customer is required.")

        if status and status not in dict(Quotation.STATUS_CHOICES):
            errors.append("Invalid status selected.")

        if not errors:
            Quotation.objects.create(
                reference=reference,
                product=product,
                customer=customer,
                quantity=quantity,
                unit_price=unit_price,
                discount_percentage=discount_percentage or 0,
                tax_percentage=tax_percentage or 0,
                status=status or "pending",
                notes=notes or None,
            )
            messages.success(request, "Quotation created successfully!")
            return redirect("quotations_index")

        old = {
            "reference": reference,
            "product_id": product_id,
            "customer_id": customer_id,
            "quantity": quantity_raw,
            "unit_price": unit_price_raw,
            "discount_percentage": discount_raw,
            "tax_percentage": tax_raw,
            "status": status,
            "notes": notes,
        }
    else:
        old = {"quantity": "1", "status": "pending"}

    context = {
        "errors": errors,
        "old": old,
        "products": products,
        "customers": customers,
        "status_choices": Quotation.STATUS_CHOICES,
    }
    return render(request, "quotations/add.html", context)


@login_required
def quotation_edit(request, pk):
    quotation = get_object_or_404(Quotation, pk=pk)
    products = Product.objects.order_by("name")
    customers = Customer.objects.order_by("name")
    errors = []

    if request.method == "POST":
        reference = request.POST.get("reference", "").strip()
        product_id = request.POST.get("product_id")
        customer_id = request.POST.get("customer_id")
        quantity_raw = request.POST.get("quantity", "1").strip()
        unit_price_raw = request.POST.get("unit_price", "0").strip()
        discount_raw = request.POST.get("discount_percentage", "0").strip()
        tax_raw = request.POST.get("tax_percentage", "0").strip()
        status = request.POST.get("status", quotation.status).strip()
        notes = request.POST.get("notes", "").strip()

        product = Product.objects.filter(pk=product_id).first()
        customer = Customer.objects.filter(pk=customer_id).first()

        try:
            quantity = int(quantity_raw)
            if quantity <= 0:
                raise ValueError
        except ValueError:
            errors.append("Quantity must be a positive integer.")
            quantity = None

        try:
            unit_price = Decimal(unit_price_raw)
        except Exception:
            errors.append("Unit price must be a number.")
            unit_price = None

        try:
            discount_percentage = Decimal(discount_raw or 0)
        except Exception:
            errors.append("Discount must be a number.")
            discount_percentage = None

        try:
            tax_percentage = Decimal(tax_raw or 0)
        except Exception:
            errors.append("Tax must be a number.")
            tax_percentage = None

        if not reference:
            errors.append("Reference is required.")
        elif (
            Quotation.objects.filter(reference=reference)
            .exclude(pk=quotation.pk)
            .exists()
        ):
            errors.append("Reference must be unique.")

        if not product:
            errors.append("Product is required.")

        if not customer:
            errors.append("Customer is required.")

        if status and status not in dict(Quotation.STATUS_CHOICES):
            errors.append("Invalid status selected.")

        if not errors:
            quotation.reference = reference
            quotation.product = product
            quotation.customer = customer
            quotation.quantity = quantity
            quotation.unit_price = unit_price
            quotation.discount_percentage = discount_percentage or 0
            quotation.tax_percentage = tax_percentage or 0
            quotation.status = status or quotation.status
            quotation.notes = notes or None
            quotation.save()
            messages.success(request, "Quotation updated successfully!")
            return redirect("quotations_index")

        old = {
            "reference": reference,
            "product_id": product_id,
            "customer_id": customer_id,
            "quantity": quantity_raw,
            "unit_price": unit_price_raw,
            "discount_percentage": discount_raw,
            "tax_percentage": tax_raw,
            "status": status,
            "notes": notes,
        }
    else:
        old = {
            "reference": quotation.reference,
            "product_id": quotation.product_id,
            "customer_id": quotation.customer_id,
            "quantity": quotation.quantity,
            "unit_price": quotation.unit_price,
            "discount_percentage": quotation.discount_percentage,
            "tax_percentage": quotation.tax_percentage,
            "status": quotation.status,
            "notes": quotation.notes or "",
        }

    context = {
        "errors": errors,
        "old": old,
        "products": products,
        "customers": customers,
        "quotation": quotation,
        "status_choices": Quotation.STATUS_CHOICES,
    }
    return render(request, "quotations/edit.html", context)


@login_required
def quotation_delete(request, pk):
    quotation = get_object_or_404(Quotation, pk=pk)

    if request.method == "POST":
        quotation.delete()
        messages.success(request, "Quotation deleted successfully.")
        return redirect("quotations_index")

    quotation.delete()
    messages.success(request, "Quotation deleted successfully.")
    return redirect("quotations_index")


@login_required
def purchases_index(request):
    purchases = Purchase.objects.select_related("supplier", "product").order_by(
        "-purchase_date", "-created_at"
    )
    return render(request, "purchases/list.html", {"purchases": purchases})


@login_required
def purchase_create(request):
    suppliers = Supplier.objects.all().order_by("name")
    products = Product.objects.all().order_by("name")
    errors = []

    if request.method == "POST":
        supplier_name = request.POST.get("supplier_name", "").strip()
        product_name = request.POST.get("product_name", "").strip()
        purchase_date = request.POST.get("purchase_date", "").strip()
        quantity = request.POST.get("quantity", "").strip()
        unit_price = request.POST.get("unit_price", "").strip()
        discount = request.POST.get("discount", "").strip()
        tax_rate = request.POST.get("tax_rate", "").strip()
        status = request.POST.get("status", "pending").strip()
        paid_amount = request.POST.get("paid_amount", "").strip()
        description = request.POST.get("description", "").strip()

        if not supplier_name:
            errors.append("Supplier name is required.")
        if not product_name:
            errors.append("Product name is required.")
        if not purchase_date:
            errors.append("Purchase date is required.")
        if not quantity:
            errors.append("Quantity is required.")
        if not unit_price:
            errors.append("Unit price is required.")
        if discount == "":
            errors.append("Discount is required.")
        if tax_rate == "":
            errors.append("Tax rate is required.")
        if not status:
            errors.append("Status is required.")
        if paid_amount == "":
            errors.append("Paid amount is required.")

        supplier = Supplier.objects.filter(name=supplier_name).first()
        if not supplier:
            errors.append("Selected supplier does not exist.")

        product = Product.objects.filter(name=product_name).first()
        if not product:
            errors.append("Selected product does not exist.")

        try:
            quantity_val = int(quantity)
            if quantity_val < 0:
                errors.append("Quantity must be at least 0.")
        except Exception:
            errors.append("Quantity must be an integer.")
            quantity_val = 0

        def to_decimal(value, field_name, min_val=None, max_val=None):
            nonlocal errors
            try:
                d = Decimal(value)
            except Exception:
                errors.append(f"{field_name} must be a valid number.")
                return Decimal("0")
            if min_val is not None and d < min_val:
                errors.append(f"{field_name} must be at least {min_val}.")
            if max_val is not None and d > max_val:
                errors.append(f"{field_name} must not exceed {max_val}.")
            return d

        unit_price_val = to_decimal(unit_price, "Unit price", Decimal("0"))
        discount_val = to_decimal(discount, "Discount", Decimal("0"), Decimal("100"))
        tax_rate_val = to_decimal(tax_rate, "Tax rate", Decimal("0"), Decimal("100"))
        paid_amount_val = to_decimal(paid_amount, "Paid amount", Decimal("0"))

        discount_amt, tax_amount_val, line_total_val = compute_purchase_totals(
            quantity_val,
            unit_price_val,
            discount_val,
            tax_rate_val,
        )

        if paid_amount_val > line_total_val:
            errors.append("Paid amount cannot exceed the line total.")

        if paid_amount_val == 0:
            payment_status = Purchase.PaymentStatusChoices.UNPAID
        elif paid_amount_val < line_total_val:
            payment_status = Purchase.PaymentStatusChoices.PARTIAL
        else:
            payment_status = Purchase.PaymentStatusChoices.PAID

        if not errors:
            with transaction.atomic():
                Purchase.objects.create(
                    supplier=supplier,
                    product=product,
                    reference=str(Purchase.objects.count() + 1),
                    purchase_date=purchase_date,
                    quantity=quantity_val,
                    unit_price=unit_price_val,
                    discount=discount_val,
                    tax_rate=tax_rate_val,
                    tax_amount=tax_amount_val,
                    line_total=line_total_val,
                    status=status,
                    paid_amount=paid_amount_val,
                    description=description or None,
                    payment_status=payment_status,
                    is_quantity_added_to_product=False,
                )

            messages.success(request, "Purchase created successfully!")
            return redirect("purchases_index")

        old = {
            "supplier_name": supplier_name,
            "product_name": product_name,
            "purchase_date": purchase_date,
            "quantity": quantity,
            "unit_price": unit_price,
            "discount": discount,
            "tax_rate": tax_rate,
            "status": status,
            "paid_amount": paid_amount,
            "description": description,
        }
    else:
        errors = []
        old = {}

    context = {
        "errors": errors,
        "old": old,
        "suppliers": suppliers,
        "products": products,
    }
    return render(request, "purchases/add.html", context)


@login_required
def purchase_edit(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    suppliers = Supplier.objects.all().order_by("name")
    products = Product.objects.all().order_by("name")
    errors = []

    if request.method == "POST":
        supplier_name = request.POST.get("supplier_name", "").strip()
        product_name = request.POST.get("product_name", "").strip()
        purchase_date = request.POST.get("purchase_date", "").strip()
        quantity = request.POST.get("quantity", "").strip()
        unit_price = request.POST.get("unit_price", "").strip()
        discount = request.POST.get("discount", "").strip()
        tax_rate = request.POST.get("tax_rate", "").strip()
        status = request.POST.get("status", purchase.status).strip()
        paid_amount = request.POST.get("paid_amount", "").strip()
        description = request.POST.get("description", "").strip()

        if not supplier_name:
            errors.append("Supplier name is required.")
        if not product_name:
            errors.append("Product name is required.")
        if not purchase_date:
            errors.append("Purchase date is required.")
        if not quantity:
            errors.append("Quantity is required.")
        if not unit_price:
            errors.append("Unit price is required.")
        if discount == "":
            errors.append("Discount is required.")
        if tax_rate == "":
            errors.append("Tax rate is required.")
        if not status:
            errors.append("Status is required.")
        if paid_amount == "":
            errors.append("Paid amount is required.")

        supplier = Supplier.objects.filter(name=supplier_name).first()
        if not supplier:
            errors.append("Selected supplier does not exist.")
        product = Product.objects.filter(name=product_name).first()
        if not product:
            errors.append("Selected product does not exist.")

        try:
            quantity_val = int(quantity)
            if quantity_val < 0:
                errors.append("Quantity must be at least 0.")
        except Exception:
            errors.append("Quantity must be an integer.")
            quantity_val = 0

        def to_decimal(value, field_name, min_val=None, max_val=None):
            nonlocal errors
            try:
                d = Decimal(value)
            except Exception:
                errors.append(f"{field_name} must be a valid number.")
                return Decimal("0")
            if min_val is not None and d < min_val:
                errors.append(f"{field_name} must be at least {min_val}.")
            if max_val is not None and d > max_val:
                errors.append(f"{field_name} must not exceed {max_val}.")
            return d

        unit_price_val = to_decimal(unit_price, "Unit price", Decimal("0"))
        discount_val = to_decimal(discount, "Discount", Decimal("0"), Decimal("100"))
        tax_rate_val = to_decimal(tax_rate, "Tax rate", Decimal("0"), Decimal("100"))
        paid_amount_val = to_decimal(paid_amount, "Paid amount", Decimal("0"))

        discount_amt, tax_amount_val, line_total_val = compute_purchase_totals(
            quantity_val,
            unit_price_val,
            discount_val,
            tax_rate_val,
        )

        if paid_amount_val > line_total_val:
            errors.append("Paid amount cannot exceed the line total.")

        if paid_amount_val == 0:
            payment_status = Purchase.PaymentStatusChoices.UNPAID
        elif paid_amount_val < line_total_val:
            payment_status = Purchase.PaymentStatusChoices.PARTIAL
        else:
            payment_status = Purchase.PaymentStatusChoices.PAID

        if not errors:
            with transaction.atomic():
                purchase.supplier = supplier
                purchase.product = product
                purchase.purchase_date = purchase_date
                purchase.quantity = quantity_val
                purchase.unit_price = unit_price_val
                purchase.discount = discount_val
                purchase.tax_rate = tax_rate_val
                purchase.tax_amount = tax_amount_val
                purchase.line_total = line_total_val
                purchase.status = status
                purchase.paid_amount = paid_amount_val
                purchase.description = description or None
                purchase.payment_status = payment_status
                purchase.save()

            messages.success(request, "Purchase updated successfully!")
            return redirect("purchases_index")

        old = {
            "supplier_name": supplier_name,
            "product_name": product_name,
            "purchase_date": purchase_date,
            "quantity": quantity,
            "unit_price": unit_price,
            "discount": discount,
            "tax_rate": tax_rate,
            "status": status,
            "paid_amount": paid_amount,
            "description": description,
        }
    else:
        old = {
            "supplier_name": purchase.supplier.name if purchase.supplier else "",
            "product_name": purchase.product.name if purchase.product else "",
            "purchase_date": purchase.purchase_date,
            "quantity": purchase.quantity,
            "unit_price": purchase.unit_price,
            "discount": purchase.discount,
            "tax_rate": purchase.tax_rate,
            "status": purchase.status,
            "paid_amount": purchase.paid_amount,
            "description": purchase.description or "",
        }

    context = {
        "errors": errors,
        "old": old,
        "purchase": purchase,
        "suppliers": suppliers,
        "products": products,
    }
    return render(request, "purchases/edit.html", context)


@login_required
def purchase_delete(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    if request.method == "POST":
        purchase.delete()
        messages.success(request, "Purchase deleted successfully.")
        return redirect("purchases_index")
    purchase.delete()
    messages.success(request, "Purchase deleted successfully.")
    return redirect("purchases_index")


@login_required
def purchase_more_options(request, pk):
    purchase = get_object_or_404(
        Purchase.objects.select_related("product", "supplier"), pk=pk
    )
    return render(request, "purchases/more_options.html", {"purchase": purchase})


@login_required
def purchase_adjust_quantity(request):
    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("purchases_index")

    purchase_id = request.POST.get("purchase_id")
    if not purchase_id:
        messages.error(request, "Purchase ID is required.")
        return redirect("purchases_index")

    try:
        purchase_id_int = int(purchase_id)
    except Exception:
        messages.error(request, "Invalid purchase ID.")
        return redirect("purchases_index")

    with transaction.atomic():
        purchase = get_object_or_404(
            Purchase.objects.select_related("product"), pk=purchase_id_int
        )
        product = purchase.product
        if not product:
            messages.error(request, "This purchase has no product attached.")
            return redirect("purchases_index")

        if not purchase.is_quantity_added_to_product:

            product.quantity += purchase.quantity
            product.save(update_fields=["quantity"])
            purchase.is_quantity_added_to_product = True
            purchase.save(update_fields=["is_quantity_added_to_product"])
            messages.success(request, "Product quantity incremented successfully!")
        else:

            if product.quantity < purchase.quantity:
                messages.warning(request, "Not enough stock to remove this purchase!")
                return redirect("purchases_index")
            product.quantity -= purchase.quantity
            product.save(update_fields=["quantity"])
            purchase.is_quantity_added_to_product = False
            purchase.save(update_fields=["is_quantity_added_to_product"])
            messages.success(request, "Product quantity decremented successfully!")

    return redirect("purchases_index")


@login_required
def wordle_view(request):

    return render(
        request,
        "wordle.html",
    )
